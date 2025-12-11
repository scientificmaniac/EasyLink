import json
import os
from typing import List, Tuple, Any
from tkinter import messagebox
from openpyxl import load_workbook
from contact import Contact
from validator import Validator
import logging

# 配置日志
logger = logging.getLogger(__name__)

class ExcelImporter:
    @staticmethod
    def import_from_excel(file_path: str) -> Tuple[bool, List[Contact]]:
        """从Excel文件导入联系人"""
        contacts: List[Contact] = []
        try:
            if not os.path.exists(file_path):
                messagebox.showerror("错误", "文件不存在")
                return False, []
                
            if not file_path.lower().endswith('.xlsx'):
                messagebox.showerror("错误", "不是有效的Excel文件")
                return False, []
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 跳过表头
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 2:  # 至少需要姓名和电话列
                    continue
                    
                name = str(row[0]) if row[0] is not None else ""
                phone = str(row[1]) if row[1] is not None else ""
                email = str(row[2]) if row[2] is not None else ""
                # 国家/地区由系统自动根据电话生成，不使用导入值
                remark = str(row[4]) if len(row) > 4 and row[4] is not None else ""
                is_frequent = row[5] == "是" if len(row) > 5 and row[5] is not None else False
                
                # 验证联系人数据
                valid, msg = Validator.validate_contact_data(name, phone, email, remark)
                if valid:
                    contact = Contact(name, phone, email, remark, is_frequent)
                    contacts.append(contact)
                else:
                    logger.warning(f"Skipping invalid contact from Excel: {msg}")
            
            logger.info(f"Excel import: loaded {len(contacts)} valid contacts")
            return True, contacts
        except PermissionError as e:
            messagebox.showerror("错误", f"没有读取权限: {str(e)}")
            logger.error(f"Excel import permission error: {e}")
            return False, []
        except Exception as e:
            messagebox.showerror("错误", f"Excel导入失败: {str(e)}")
            logger.error(f"Excel import failed: {e}", exc_info=True)
            return False, []

class TXTImporter:
    @staticmethod
    def import_from_txt(file_path: str) -> Tuple[bool, List[Contact]]:
        """从TXT文件导入联系人"""
        contacts: List[Contact] = []
        try:
            if not os.path.exists(file_path):
                messagebox.showerror("错误", "文件不存在")
                return False, []
                
            if not file_path.lower().endswith('.txt'):
                messagebox.showerror("错误", "不是有效的TXT文件")
                return False, []
            
            with open(file_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            
            current_contact = {}
            for line in lines:
                line = line.strip()
                if line.startswith("联系人 ") and line.endswith(":"):
                    if current_contact and "name" in current_contact and "phone" in current_contact:
                        # 保存上一个联系人
                        try:
                            contact = Contact(
                                current_contact["name"],
                                current_contact["phone"],
                                current_contact.get("email", ""),
                                current_contact.get("remark", ""),
                                current_contact.get("is_frequent", False)
                            )
                            contacts.append(contact)
                        except ValueError as e:
                            logger.warning(f"Skipping invalid contact from TXT: {e}")
                        finally:
                            current_contact = {}
                elif line.startswith("姓名: "):
                    current_contact["name"] = line.replace("姓名: ", "")
                elif line.startswith("电话: "):
                    current_contact["phone"] = line.replace("电话: ", "")
                elif line.startswith("邮箱: "):
                    current_contact["email"] = line.replace("邮箱: ", "")
                elif line.startswith("备注: "):
                    current_contact["remark"] = line.replace("备注: ", "")
                elif line.startswith("常用联系人: "):
                    current_contact["is_frequent"] = line.replace("常用联系人: ", "") == "是"
            
            # 保存最后一个联系人
            if current_contact and "name" in current_contact and "phone" in current_contact:
                try:
                    contact = Contact(
                        current_contact["name"],
                        current_contact["phone"],
                        current_contact.get("email", ""),
                        current_contact.get("remark", ""),
                        current_contact.get("is_frequent", False)
                    )
                    contacts.append(contact)
                except ValueError as e:
                    logger.warning(f"Skipping invalid contact from TXT: {e}")
            
            logger.info(f"TXT import: loaded {len(contacts)} valid contacts")
            return True, contacts
        except PermissionError as e:
            messagebox.showerror("错误", f"没有读取权限: {str(e)}")
            logger.error(f"TXT import permission error: {e}")
            return False, []
        except Exception as e:
            messagebox.showerror("错误", f"TXT导入失败: {str(e)}")
            logger.error(f"TXT import failed: {e}", exc_info=True)
            return False, []

class MDImporter:
    @staticmethod
    def import_from_md(file_path: str) -> Tuple[bool, List[Contact]]:
        """从Markdown文件导入联系人"""
        contacts: List[Contact] = []
        try:
            if not os.path.exists(file_path):
                messagebox.showerror("错误", "文件不存在")
                return False, []
                
            if not file_path.lower().endswith('.md'):
                messagebox.showerror("错误", "不是有效的Markdown文件")
                return False, []
            
            with open(file_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            
            # 跳过标题和表头
            start_import = False
            for line in lines:
                line = line.strip()
                if line.startswith("|------|------|------|------------|------|------------|"):
                    start_import = True
                    continue
                
                if start_import and line.startswith("|"):
                    # 解析Markdown表格行
                    parts = [part.strip() for part in line.split("|") if part.strip()]
                    if len(parts) >= 6:
                        name = parts[0]
                        phone = parts[1]
                        email = parts[2]
                        remark = parts[4]
                        is_frequent = parts[5] == "是"
                        
                        # 验证联系人数据
                        valid, msg = Validator.validate_contact_data(name, phone, email, remark)
                        if valid:
                            contact = Contact(name, phone, email, remark, is_frequent)
                            contacts.append(contact)
                        else:
                            logger.warning(f"Skipping invalid contact from MD: {msg}")
            
            logger.info(f"Markdown import: loaded {len(contacts)} valid contacts")
            return True, contacts
        except PermissionError as e:
            messagebox.showerror("错误", f"没有读取权限: {str(e)}")
            logger.error(f"Markdown import permission error: {e}")
            return False, []
        except Exception as e:
            messagebox.showerror("错误", f"Markdown导入失败: {str(e)}")
            logger.error(f"Markdown import failed: {e}", exc_info=True)
            return False, []

class JSONImporter:
    @staticmethod
    def import_from_json(file_path: str) -> Tuple[bool, List[Contact]]:
        """从JSON文件导入联系人"""
        contacts: List[Contact] = []
        try:
            if not os.path.exists(file_path):
                messagebox.showerror("错误", "文件不存在")
                return False, []
                
            if not file_path.lower().endswith('.json'):
                messagebox.showerror("错误", "不是有效的JSON文件")
                return False, []
            
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            if not isinstance(data, list):
                messagebox.showerror("错误", "JSON数据格式不正确，应为联系人列表")
                return False, []
            
            for contact_data in data:
                if isinstance(contact_data, dict) and "name" in contact_data and "phone" in contact_data:
                    name = str(contact_data["name"]) if contact_data["name"] is not None else ""
                    phone = str(contact_data["phone"]) if contact_data["phone"] is not None else ""
                    email = str(contact_data.get("email", "")) if contact_data.get("email") is not None else ""
                    remark = str(contact_data.get("remark", "")) if contact_data.get("remark") is not None else ""
                    is_frequent = bool(contact_data.get("is_frequent", False))
                    
                    # 验证联系人数据
                    valid, msg = Validator.validate_contact_data(name, phone, email, remark)
                    if valid:
                        contact = Contact(name, phone, email, remark, is_frequent)
                        contacts.append(contact)
                    else:
                        logger.warning(f"Skipping invalid contact from JSON: {msg}")
            
            logger.info(f"JSON import: loaded {len(contacts)} valid contacts")
            return True, contacts
        except json.JSONDecodeError as e:
            messagebox.showerror("错误", f"JSON格式错误: {str(e)}")
            logger.error(f"JSON import decode error: {e}")
            return False, []
        except PermissionError as e:
            messagebox.showerror("错误", f"没有读取权限: {str(e)}")
            logger.error(f"JSON import permission error: {e}")
            return False, []
        except Exception as e:
            messagebox.showerror("错误", f"JSON导入失败: {str(e)}")
            logger.error(f"JSON import failed: {e}", exc_info=True)
            return False, []

class DataImporter:
    @staticmethod
    def import_contacts(file_path: str, contact_manager) -> bool:
        """统一导入入口，根据文件扩展名自动选择导入方式"""
        if not isinstance(file_path, str):
            messagebox.showerror("错误", "文件路径必须是字符串")
            return False
        
        if not os.path.exists(file_path):
            messagebox.showerror("错误", "文件不存在")
            return False
        
        if not os.path.isfile(file_path):
            messagebox.showerror("错误", "不是有效的文件")
            return False
        
        ext = os.path.splitext(file_path)[1].lower()
        success = False
        contacts: List[Contact] = []
        
        if ext == ".xlsx":
            success, contacts = ExcelImporter.import_from_excel(file_path)
        elif ext == ".txt":
            success, contacts = TXTImporter.import_from_txt(file_path)
        elif ext == ".md":
            success, contacts = MDImporter.import_from_md(file_path)
        elif ext == ".json":
            success, contacts = JSONImporter.import_from_json(file_path)
        else:
            messagebox.showerror("错误", "不支持的文件格式")
            return False
        
        if success and contacts:
            # 导入联系人到系统
            imported_count = 0
            duplicate_count = 0
            
            for contact in contacts:
                # 检查电话号码是否已存在
                phone_exists = False
                for existing_contact in contact_manager.storage.contacts:
                    if existing_contact.phone == contact.phone:
                        phone_exists = True
                        break
                
                if not phone_exists:
                    result, msg = contact_manager.add_contact(contact)
                    if result:
                        imported_count += 1
                    else:
                        logger.warning(f"Failed to add imported contact: {msg}")
                        duplicate_count += 1
                else:
                    duplicate_count += 1
            
            messagebox.showinfo(
                "导入完成", 
                f"成功导入 {imported_count} 个联系人\n" 
                f"跳过 {duplicate_count} 个重复或无效联系人\n" 
                f"共处理 {len(contacts)} 个联系人"
            )
            logger.info(f"Import completed: {imported_count} contacts imported, {duplicate_count} skipped")
            return True
        else:
            messagebox.showinfo("导入结果", "未导入任何联系人")
            return False